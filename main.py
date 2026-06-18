from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from groq import Groq
import os
import io
from dotenv import load_dotenv
import pymupdf
import docx

# Load environment variables
load_dotenv()

# Initialize FastAPI app
app = FastAPI(
    title="Project Estimation API",
    description="AI-powered project estimation engine",
    version="1.0.0"
)

# Allow frontend to call this API
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Initialize Groq client
GROQ_API_KEY = os.getenv("GROQ_API_KEY")

# Request model
class EstimationRequest(BaseModel):
    requirements: str
    hourly_rate: float = 50.0
    currency: str = "USD"

# The core estimation prompt
def build_prompt(requirements: str, hourly_rate: float, currency: str) -> str:
    return f"""
You are a senior software project estimation expert with 20 years of IT delivery experience.

Analyze the following requirements and provide a detailed estimation.

Respond in this EXACT format:

## SCOPE SUMMARY
[3 sentences describing what will be built]

## WORK BREAKDOWN STRUCTURE
[List every work package as: WP1: Name - Description]

## EFFORT ESTIMATE
[For each WP: WP1: X-Y hours]

## TEAM RECOMMENDATION
[List roles and count needed]

## TIMELINE
[Total weeks to deliver with phases]

## TOP 5 RISKS
[Risk 1: Name - Description - Mitigation]

## TOTAL COST RANGE
[Min cost - Max cost in {currency} at {hourly_rate}/hour]

## ASSUMPTIONS
[List all assumptions made]

Requirements:
{requirements}
"""

# Route 1 — Estimate from text
@app.post("/estimate")
async def estimate_from_text(request: EstimationRequest):
    if not request.requirements.strip():
        raise HTTPException(status_code=400, detail="Requirements cannot be empty")
    
    if len(request.requirements) < 20:
        raise HTTPException(status_code=400, detail="Requirements too short. Please provide more detail.")

    try:
        prompt = build_prompt(
            request.requirements,
            request.hourly_rate,
            request.currency
        )

        response = client.chat.completions.create(
            model="llama-3.3-70b-versatile",
            messages=[
                {
                    "role": "system",
                    "content": "You are a senior software project estimation expert. Always respond in the exact format requested. Be specific with numbers and ranges."
                },
                {
                    "role": "user",
                    "content": prompt
                }
            ],
            temperature=0.3
        )

        result = response.choices[0].message.content

        return {
            "status": "success",
            "estimation": result,
            "requirements_length": len(request.requirements),
            "model_used": "llama-3.3-70b-versatile"
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"AI Error: {str(e)}")


# Route 2 — Estimate from uploaded PDF or Word file
@app.post("/estimate/upload")
async def estimate_from_file(
    file: UploadFile = File(...),
    hourly_rate: float = 50.0,
    currency: str = "USD"
):
    # Check file type
    allowed_types = ["application/pdf", 
                     "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                     "text/plain"]
    
    filename = file.filename.lower()
    
    if not (filename.endswith(".pdf") or 
            filename.endswith(".docx") or 
            filename.endswith(".txt")):
        raise HTTPException(
            status_code=400, 
            detail="Only PDF, DOCX, and TXT files are supported"
        )

    try:
        contents = await file.read()
        extracted_text = ""

        # Extract text from PDF
        if filename.endswith(".pdf"):
            pdf_document = pymupdf.open(stream=contents, filetype="pdf")
            for page in pdf_document:
                extracted_text += page.get_text()
            pdf_document.close()

        # Extract text from Word document
        elif filename.endswith(".docx"):
            doc = docx.Document(io.BytesIO(contents))
            for paragraph in doc.paragraphs:
                extracted_text += paragraph.text + "\n"

        # Extract text from plain text file
        elif filename.endswith(".txt"):
            extracted_text = contents.decode("utf-8")

        if not extracted_text.strip():
            raise HTTPException(
                status_code=400, 
                detail="Could not extract text from file. File may be empty or image-based."
            )

        # Now estimate using extracted text
        prompt = build_prompt(extracted_text, hourly_rate, currency)

        response = client.chat.completions.create(
            model="llama-3.3-70b-versatile",
            messages=[
                {
                    "role": "system",
                    "content": "You are a senior software project estimation expert. Always respond in the exact format requested."
                },
                {
                    "role": "user",
                    "content": prompt
                }
            ],
            temperature=0.3
        )

        result = response.choices[0].message.content

        return {
            "status": "success",
            "filename": file.filename,
            "extracted_text_length": len(extracted_text),
            "estimation": result,
            "model_used": "llama-3.3-70b-versatile"
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


# Route 3 — Health check
@app.get("/")
async def root():
    return {
        "status": "running",
        "message": "Project Estimation API is live",
        "version": "1.0.0"
    }


# Route 4 — Health check endpoint
@app.get("/health")
async def health():
    return {"status": "healthy"}

# Route 5 
from fastapi.responses import StreamingResponse
import io
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
from reportlab.lib.units import inch
from reportlab.lib import colors
from pydantic import BaseModel as PydanticBaseModel

class ExportRequest(PydanticBaseModel):
    estimation: str
    project_name: str = "Project Estimation"

@app.post("/export/pdf")
async def export_pdf(request: ExportRequest):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=inch,
        leftMargin=inch,
        topMargin=inch,
        bottomMargin=inch
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Title'],
        fontSize=24,
        textColor=colors.HexColor('#2563eb'),
        spaceAfter=20
    )
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#1e40af'),
        spaceBefore=16,
        spaceAfter=8,
        borderPad=4
    )
    body_style = ParagraphStyle(
        'CustomBody',
        parent=styles['Normal'],
        fontSize=11,
        leading=16,
        spaceAfter=6
    )
    story = []
    story.append(Paragraph("EstimateAI", title_style))
    story.append(Paragraph(request.project_name, styles['Heading1']))
    story.append(Spacer(1, 20))
    lines = request.estimation.split('\n')
    for line in lines:
        if line.startswith('## '):
            story.append(Spacer(1, 8))
            story.append(Paragraph(line.replace('## ', ''), heading_style))
        elif line.startswith('- ') or line.startswith('* '):
            story.append(Paragraph(
                '• ' + line.replace('- ', '').replace('* ', ''),
                body_style
            ))
        elif line.strip():
            story.append(Paragraph(line, body_style))
    doc.build(story)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename=estimation.pdf"
        }
    )
# rRoute
class ExcelExportRequest(BaseModel):
    estimation: str
    project_name: str = "Project Estimation"

@app.post("/export/excel")
async def export_excel(request: ExcelExportRequest):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Estimation"

    ws.merge_cells('A1:D1')
    ws['A1'] = f"EstimateAI — {request.project_name}"
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill("solid", fgColor="1e3a8a")
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20

    row = 3
    lines = request.estimation.split('\n')

    for line in lines:
        if line.startswith('## '):
            ws.merge_cells(f'A{row}:D{row}')
            ws[f'A{row}'] = line.replace('## ', '').strip()
            ws[f'A{row}'].font = Font(size=12, bold=True, color="FFFFFF")
            ws[f'A{row}'].fill = PatternFill("solid", fgColor="2563eb")
            ws[f'A{row}'].alignment = Alignment(vertical="center")
            ws.row_dimensions[row].height = 28
            row += 1
        elif line.strip():
            ws.merge_cells(f'A{row}:D{row}')
            ws[f'A{row}'] = line.strip()
            ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[row].height = 20
            row += 1

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={request.project_name}.xlsx"}
    )